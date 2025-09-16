import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from rest_framework.parsers import MultiPartParser, FormParser
from core_apps.attendance.models import AttendanceRecord
from django.core.paginator import Paginator
import pandas as pd
import numpy as np
from django.db.models import Sum, Count, F, ExpressionWrapper, FloatField, Avg
from datetime import timedelta
from django.utils import timezone
from loguru import logger
from django.core.paginator import Paginator

from .serializers import WorkbookUploadSerializer
from .models import UploadedWorkbook
from core_apps.common.models import Archdeaconry, Parish, Congregation
from core_apps.attendance.models import AttendanceRecord

class WorkbookUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        serializer = WorkbookUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        files = serializer.validated_data['files']

        summary = []

        for file in files:
            file_name = file.name
            file_summary = {
                'file': file_name,
                'new_upload': False,
                'processed': False,
                'errors': []
            }

            # Get or create the UploadedWorkbook record
            upload, created = UploadedWorkbook.objects.get_or_create(
                file_name=file_name,
                defaults={'file': file}
            )
            file_summary['new_upload'] = created

            if not created:
                upload.file = file
                upload.processed = False
                upload.save()

            # Try to open workbook
            try:
                wb = openpyxl.load_workbook(file)
            except Exception as e:
                file_summary['errors'].append(f"Failed to read file: {str(e)}")
                summary.append(file_summary)
                continue

            sheet_date = None

            # Process each sheet
            for sheet_name in wb.sheetnames:
                try:
                    sheet_date = datetime.datetime.strptime(sheet_name, '%d-%m-%y').date()
                except ValueError:
                    file_summary['errors'].append(f"Invalid sheet name format: '{sheet_name}'")
                    continue

                ws = wb[sheet_name]

                # Process each row
                for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    try:
                        arch_name = (row[0].value or '').strip()
                        parish_name = (row[1].value or '').strip()
                        cong_name = (row[2].value or '').strip()
                        if not (arch_name and parish_name and cong_name):
                            continue

                        ss = row[3].value or 0
                        youth = row[4].value or 0
                        adults = row[5].value or 0
                        diff_abled = row[6].value or 0
                        total_col = row[8].value or 0
                        banked = row[9].value or 0
                        unbanked = row[10].value or 0
                        remarks = row[11].value or ''

                        arch, _ = Archdeaconry.objects.get_or_create(name=arch_name)
                        parish, _ = Parish.objects.get_or_create(name=parish_name, archdeaconry=arch)
                        cong, _ = Congregation.objects.get_or_create(name=cong_name, parish=parish)

                        AttendanceRecord.objects.update_or_create(
                            workbook=upload,
                            archdeaconry=arch,
                            parish=parish,
                            congregation=cong,
                            sunday_date=sheet_date,
                            defaults={
                                'sunday_school': ss,
                                'adults': adults,
                                'diff_abled': diff_abled,
                                'youth': youth,
                                'total_collection': total_col,
                                'banked': banked,
                                'unbanked': unbanked,
                                'remarks': remarks,
                            }
                        )

                    except Exception as row_err:
                        file_summary['errors'].append(
                            f"Sheet '{sheet_name}', row {idx}: {str(row_err)}"
                        )
                        # continue processing the rest of the rows

            # Finalize upload record
            upload.sheet_date = sheet_date or timezone.now().date()
            # Only set upload.processed=True if there were no errors
            upload.processed = len(file_summary['errors']) == 0
            upload.save()

            # Mark processed in summary only if no errors
            file_summary['processed'] = upload.processed
            summary.append(file_summary)

        return Response(summary, status=status.HTTP_200_OK)

class DashboardAnalytics(APIView):
    def get(self, request):
        try:
            start_date_str = request.query_params.get('start_date')
            end_date_str = request.query_params.get('end_date')

            end_date = pd.to_datetime(end_date_str).date() if end_date_str else timezone.now().date()
            start_date = pd.to_datetime(start_date_str).date() if start_date_str else (end_date - timedelta(days=730))

            # Optional filters
            archdeaconry_id = request.query_params.get('archdeaconry')
            parish_id = request.query_params.get('parish')
            congregation_id = request.query_params.get('congregation')

            # Base queryset
            records = AttendanceRecord.objects.filter(sunday_date__range=(start_date, end_date)).select_related(
                'congregation__parish__archdeaconry'
            )

            if congregation_id:
                records = records.filter(congregation_id=congregation_id)
            elif parish_id:
                records = records.filter(congregation__parish_id=parish_id)
            elif archdeaconry_id:
                records = records.filter(congregation__parish__archdeaconry_id=archdeaconry_id)

            if not records.exists():
                return Response({"detail": "No data available"}, status=404)
            
            # Convert to DataFrame with proper typing
            df = pd.DataFrame.from_records(records.values(
                'id','sunday_date', 'sunday_school', 'adults', 'youth', 'diff_abled',
                'total_collection', 'banked', 'unbanked',
                congregation_name=F('congregation__name'),
                parish_name=F('congregation__parish__name'),
                archdeaconry_name=F('congregation__parish__archdeaconry__name')
            ))
            
            # Ensure numeric columns
            numeric_cols = ['sunday_school', 'adults', 'youth', 'diff_abled', 'total_collection', 'banked', 'unbanked']
            df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
            
            # Calculate metrics
            df['total_attendance'] = df[['sunday_school', 'adults', 'youth', 'diff_abled']].sum(axis=1)
            df['month'] = pd.to_datetime(df['sunday_date']).dt.to_period('M')
            
            # 1. Overall statistics
            overall_stats = {
                "total_total_collection": float(df['total_collection'].sum()),
                "avg_weekly_attendance": float(df['total_attendance'].mean()),
                "growth_rate": self.calculate_growth_rate(df),
                "banked_percentage": float((df['banked'].sum() / df['total_collection'].sum()) * 100),
            }
            
            # 2. Time series analysis
            time_series = df.groupby('month').agg({
                'total_attendance': 'mean',
                'total_collection': 'sum',
                'sunday_school': 'mean',
                'adults': 'mean'
            }).reset_index()
            time_series['month'] = time_series['month'].astype(str)
             # Format numbers to prevent serialization issues
            time_series['total_attendance'] = time_series['total_attendance'].round(1)
            time_series['total_collection'] = time_series['total_collection'].round(2)
            time_series = time_series.to_dict(orient='records')
            
            # 3. Hierarchy analysis
            hierarchy_stats = df.groupby(['archdeaconry_name', 'parish_name', 'congregation_name']).agg({
                'total_attendance': 'mean',
                'total_collection': 'sum',
                'id': 'count'
            }).reset_index().to_dict(orient='records')
            
            # 4. Financial analysis - with fixed nlargest
            financial_analysis = {
                "banked_percentage": float((df['banked'].sum() / df['total_collection'].sum()) * 100),
                "top_congregations": df.groupby('congregation_name')['total_collection']
                                    .sum()
                                    .sort_values(ascending=False)
                                    .head(5)
                                    .to_dict(),
                "collection_trend": self.calculate_collection_trend(df)
            }
            
            # 5. Attendance segmentation
            attendance_segmentation = {
                "sunday_school_avg": float(df['sunday_school'].mean()),
                "adults_avg": float(df['adults'].mean()),
                "youth_avg": float(df['youth'].mean()),
                "diff_abled_avg": float(df['diff_abled'].mean()),
                "segmentation_ratio": {
                    "sunday_school": float((df['sunday_school'].sum() / df['total_attendance'].sum()) * 100),
                    "adults": float((df['adults'].sum() / df['total_attendance'].sum()) * 100),
                    "youth": float((df['youth'].sum() / df['total_attendance'].sum()) * 100),
                    "diff_abled": float((df['diff_abled'].sum() / df['total_attendance'].sum()) * 100),
                }
            }
            
            return Response({
                "overall": overall_stats,
                "time_series": time_series,
                "hierarchy": hierarchy_stats,
                "financial": financial_analysis,
                "attendance_segmentation": attendance_segmentation
            })
            
        except Exception as e:
            logger.exception("Error in dashboard analytics")
            return Response(
                {"error": "An error occurred while processing analytics data"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def calculate_growth_rate(self, df):
        df_sorted = df.sort_values(by='sunday_date')
        if df_sorted.empty:
            return 0.0
        first = df_sorted.iloc[0]['total_attendance']
        last = df_sorted.iloc[-1]['total_attendance']
        if first == 0:
            return 0.0
        growth_rate = ((last - first) / first) * 100
        return round(growth_rate, 2)

    def calculate_collection_trend(self, df):
        monthly = df.groupby('month')['total_collection'].sum().reset_index()
        monthly['month'] = monthly['month'].astype(str)
        return monthly.to_dict(orient='records')


# views.py
class ArchdeaconryListView(APIView):
    def get(self, request):
        archdeaconries = Archdeaconry.objects.all().values('id', 'name')
        return Response(archdeaconries)

class ParishListView(APIView):
    def get(self, request):
        archdeaconry_id = request.query_params.get('archdeaconry_id')
        queryset = Parish.objects.all()
        
        if archdeaconry_id:
            queryset = queryset.filter(archdeaconry_id=archdeaconry_id)
            
        parishes = queryset.values('id', 'name', 'archdeaconry_id')
        return Response(parishes)

class CongregationListView(APIView):
    def get(self, request):
        parish_id = request.query_params.get('parish_id')
        queryset = Congregation.objects.all()
        
        if parish_id:
            queryset = queryset.filter(parish_id=parish_id)
            
        congregations = queryset.values('id', 'name', 'parish_id')
        return Response(congregations)
    
class CongregationsByArchdeaconryView(APIView):
    def get(self, request):
        parish_ids = request.query_params.get('parish_ids', '').split(',')
        parish_ids = [id for id in parish_ids if id]  # Remove empty strings
        
        if not parish_ids:
            return Response([])
            
        congregations = Congregation.objects.filter(
            parish_id__in=parish_ids
        ).values('id', 'name', 'parish_id')
        
        return Response(congregations)

class RecordsListView(APIView):
    def get(self, request):
        # Get filters from request
        archdeaconry_id = request.query_params.get('archdeaconry')
        parish_id = request.query_params.get('parish')
        congregation_id = request.query_params.get('congregation')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('pageSize', 10))

        # Build query
        queryset = AttendanceRecord.objects.select_related(
            'congregation__parish__archdeaconry'
        ).annotate(
            archdeaconry_name=F('congregation__parish__archdeaconry__name'),
            parish_name=F('congregation__parish__name'),
            congregation_name=F('congregation__name'),
            total_attendance=F('sunday_school') + F('adults') + F('youth') + F('diff_abled')
        ).order_by('-sunday_date')
        
        # Apply filters
        if congregation_id:
            queryset = queryset.filter(congregation_id=congregation_id)
        elif parish_id:
            queryset = queryset.filter(congregation__parish_id=parish_id)
        elif archdeaconry_id:
            queryset = queryset.filter(congregation__parish__archdeaconry_id=archdeaconry_id)
            
        if start_date:
            queryset = queryset.filter(sunday_date__gte=start_date)
            
        if end_date:
            queryset = queryset.filter(sunday_date__lte=end_date)
        
        # Get total count before pagination
        total_records = queryset.count()
        
        # Paginate results
        paginator = Paginator(queryset, page_size)
        page_obj = paginator.get_page(page)

        # Convert to list of dictionaries
        records = list(page_obj.object_list.values(
            'id', 'sunday_date', 'sunday_school', 'adults', 'youth', 'remarks', 'diff_abled',
            'total_collection', 'banked', 'unbanked', 'total_attendance',
            'archdeaconry_name', 'parish_name', 'congregation_name'
        ))
        
        return Response({
            'data': records,
            'total': total_records,
            'page': page,
            'pageSize': page_size,
            'totalPages': paginator.num_pages
        })
